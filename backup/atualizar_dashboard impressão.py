#!/usr/bin/env python3
"""
Dashboard de Manutenção – Atualizador com publicação automática no GitHub
=========================================================================
Uso:
    python atualizar_dashboard.py planilha.xlsx [--repo usuario/repo] [--token TOKEN]

Variáveis de ambiente (alternativa aos argumentos):
    GITHUB_TOKEN   – Personal Access Token com permissão contents:write
    GITHUB_REPO    – Ex: "joaosilva/dashboard-manutencao"
    DASHBOARD_FILE – Caminho local do HTML (padrão: dashboard_manutencao.html)

Fluxo:
    1. Lê a planilha xlsx (máquinas + compressores)
    2. Atualiza DEFAULT_DATA e COMP_DATA no HTML
    3. Salva o HTML localmente
    4. Faz commit + push do HTML para o branch gh-pages (GitHub Pages)
"""

import re
import sys
import json
import os
import argparse
import base64
import urllib.request
import urllib.error
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("❌  Instale o openpyxl:  pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────────────────────────────────────
# 1. LEITURA DA PLANILHA
# ──────────────────────────────────────────────────────────────────────────────

MACHINE_SHEETS = [
    ("Resina 1 - Pedro",       "Pedro"),
    ("Resina 2 - Neto",        "Neto"),
    ("Resina 3 - Antonio José","Antonio José"),
    ("Gel Coat 1 - PC",        "PC"),
    ("Gel Coat 2 - Dionísio",  "Dionísio"),
    ("Gel Coat 3 - Evandro",   "Evandro"),
]

COMPRESSOR_SHEETS = [
    "Compressor – Atlas Copco",
    "Compressor – Schulz SRP20",
    "Compressor – Schulz SRP15 #1",
    "Compressor – Schulz SRP15 #2",
]


def cell_str(cell):
    """Retorna valor da célula como string limpa, ou ''."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def parse_time(val):
    """Converte string de tempo (minutos) para int."""
    try:
        return int(float(str(val).replace(",", ".").strip()))
    except (ValueError, AttributeError):
        return 0


def read_machine_sheet(ws, operator):
    """Lê uma aba de máquina e retorna lista de records."""
    records = []
    for row in ws.iter_rows(min_row=1):
        date_val  = cell_str(row[0])
        problem   = cell_str(row[1]) if len(row) > 1 else ""
        parts_raw = cell_str(row[2]) if len(row) > 2 else ""
        time_raw  = cell_str(row[3]) if len(row) > 3 else ""

        if not date_val:
            continue

        time_min = parse_time(time_raw) if time_raw else 0
        is_ok    = problem.upper() in ("OK", "")
        parts_arr = [p.strip() for p in parts_raw.split(",") if p.strip()] if parts_raw else []

        records.append({
            "date":     date_val,
            "problem":  problem,
            "partsRaw": parts_raw,
            "partsArr": parts_arr,
            "timeMin":  time_min,
            "isOk":     is_ok,
            "operator": operator,
        })
    return records


def read_compressor_sheet(ws, name):
    """Lê uma aba de compressor e retorna dict no formato COMP_DATA."""
    records = []
    for row in ws.iter_rows(min_row=2):          # pula cabeçalho
        date_val    = cell_str(row[0])
        horimeter   = cell_str(row[1]) if len(row) > 1 else ""
        maintenance = cell_str(row[2]) if len(row) > 2 else ""
        parts       = cell_str(row[3]) if len(row) > 3 else ""

        if not date_val:
            continue

        records.append({
            "date":        date_val,
            "horimeter":   horimeter,
            "maintenance": maintenance,
            "parts":       parts,
        })

    # Mantém schedule como lista vazia (não vem da planilha)
    return {
        "name":     name,
        "records":  records,
        "schedule": [],
    }


def load_xlsx(path):
    """Carrega a planilha e retorna (machine_data, comp_data)."""
    print(f"📂  Lendo planilha: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames

    # ── Máquinas ──────────────────────────────────────────────────────────────
    machine_data = []
    for sheet_name, operator in MACHINE_SHEETS:
        # Busca tolerante (sem distinção maiús/minús e pequenas variações)
        matched = next(
            (s for s in sheet_names if s.strip().lower() == sheet_name.lower()),
            None
        )
        if not matched:
            print(f"  ⚠️  Aba não encontrada: '{sheet_name}' — pulando.")
            continue
        ws = wb[matched]
        records = read_machine_sheet(ws, operator)
        machine_data.append({
            "name":     sheet_name,
            "operator": operator,
            "records":  records,
        })
        print(f"  ✅  {sheet_name}: {len(records)} registros")

    # ── Compressores ──────────────────────────────────────────────────────────
    comp_data = []
    for sheet_name in COMPRESSOR_SHEETS:
        matched = next(
            (s for s in sheet_names if s.strip().lower() == sheet_name.lower()),
            None
        )
        if not matched:
            print(f"  ⚠️  Aba não encontrada: '{sheet_name}' — pulando.")
            continue
        ws = wb[matched]
        comp = read_compressor_sheet(ws, sheet_name)
        comp_data.append(comp)
        print(f"  ✅  {sheet_name}: {len(comp['records'])} registros")

    return machine_data, comp_data


# ──────────────────────────────────────────────────────────────────────────────
# 2. ATUALIZAÇÃO DO HTML
# ──────────────────────────────────────────────────────────────────────────────

def update_html(html_path, machine_data, comp_data, output_path=None):
    """Substitui DEFAULT_DATA e COMP_DATA no HTML e salva."""
    output_path = output_path or html_path

    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()

    # ── DEFAULT_DATA ──────────────────────────────────────────────────────────
    new_default = "const DEFAULT_DATA = " + json.dumps(
        machine_data, ensure_ascii=False, separators=(",", ":")
    )
    content, n1 = re.subn(
        r"const DEFAULT_DATA\s*=\s*\[.*?\](?=\s*[;\n])",
        new_default,
        content,
        count=1,
        flags=re.DOTALL,
    )
    if n1 == 0:
        print("  ⚠️  DEFAULT_DATA não encontrado no HTML — verifique o arquivo.")
    else:
        print("  ✅  DEFAULT_DATA atualizado.")

    # ── COMP_DATA ─────────────────────────────────────────────────────────────
    if comp_data:
        new_comp = "const COMP_DATA = " + json.dumps(
            comp_data, ensure_ascii=False, separators=(",", ":")
        )
        content, n2 = re.subn(
            r"const COMP_DATA\s*=\s*\[.*?\](?=\s*[;\n])",
            new_comp,
            content,
            count=1,
            flags=re.DOTALL,
        )
        if n2 == 0:
            print("  ⚠️  COMP_DATA não encontrado no HTML — ignorando compressores.")
        else:
            print("  ✅  COMP_DATA atualizado.")

    # Adiciona comentário de atualização no <head>
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    content = content.replace(
        "<!-- LAST_UPDATE -->",
        f"<!-- LAST_UPDATE: {ts} -->",
        1,
    )
    # Se não tinha a marca, insere no início do <head>
    if f"<!-- LAST_UPDATE: {ts} -->" not in content:
        content = content.replace(
            "<head>",
            f"<head>\n<!-- Atualizado automaticamente em {ts} -->",
            1,
        )

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"  💾  HTML salvo em: {output_path}")
    return output_path


# ──────────────────────────────────────────────────────────────────────────────
# 3. PUBLICAÇÃO NO GITHUB
# ──────────────────────────────────────────────────────────────────────────────

GITHUB_API = "https://api.github.com"


def github_request(method, endpoint, token, data=None):
    """Faz uma requisição à API do GitHub e retorna o JSON."""
    url = f"{GITHUB_API}{endpoint}"
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(
        url,
        data=body,
        method=method,
        headers={
            "Authorization": f"token {token}",
            "Accept":        "application/vnd.github+json",
            "Content-Type":  "application/json",
            "User-Agent":    "dashboard-updater/1.0",
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"  ❌  GitHub API error {e.code}: {body}")
        raise


def get_file_sha(repo, path, token, branch="gh-pages"):
    """Retorna o SHA do arquivo se existir, ou None."""
    try:
        data = github_request(
            "GET", f"/repos/{repo}/contents/{path}?ref={branch}", token
        )
        return data.get("sha")
    except urllib.error.HTTPError:
        return None


def ensure_branch_exists(repo, token, branch="gh-pages"):
    """Cria o branch gh-pages se não existir."""
    try:
        github_request("GET", f"/repos/{repo}/git/ref/heads/{branch}", token)
        print(f"  ℹ️  Branch '{branch}' já existe.")
        return
    except urllib.error.HTTPError:
        pass  # branch não existe, vamos criar

    # Pega o SHA do HEAD do branch padrão
    repo_info = github_request("GET", f"/repos/{repo}", token)
    default_branch = repo_info.get("default_branch", "main")
    ref_data = github_request(
        "GET", f"/repos/{repo}/git/ref/heads/{default_branch}", token
    )
    sha = ref_data["object"]["sha"]

    # Cria o novo branch
    github_request("POST", f"/repos/{repo}/git/refs", token, {
        "ref": f"refs/heads/{branch}",
        "sha": sha,
    })
    print(f"  ✅  Branch '{branch}' criado a partir de '{default_branch}'.")


def publish_to_github(html_path, repo, token, branch="gh-pages", remote_filename="index.html"):
    """Faz upload/atualização do HTML no GitHub e habilita Pages se necessário."""
    print(f"\n🚀  Publicando no GitHub ({repo} / {branch}) …")

    # Verifica/cria branch
    ensure_branch_exists(repo, token, branch)

    # Lê o arquivo local
    with open(html_path, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode()

    # Verifica se já existe (para obter SHA)
    sha = get_file_sha(repo, remote_filename, token, branch)

    commit_msg = f"chore: atualiza dashboard [{datetime.now().strftime('%d/%m/%Y %H:%M')}]"
    payload = {
        "message": commit_msg,
        "content": content_b64,
        "branch":  branch,
    }
    if sha:
        payload["sha"] = sha
        print(f"  🔄  Atualizando '{remote_filename}' (sha: {sha[:7]}…)")
    else:
        print(f"  ➕  Criando '{remote_filename}' no branch '{branch}'")

    result = github_request(
        "PUT", f"/repos/{repo}/contents/{remote_filename}", token, payload
    )

    commit_url = result.get("commit", {}).get("html_url", "")
    print(f"  ✅  Commit: {commit_url}")

    # Tenta habilitar GitHub Pages (ignora erro se já estiver habilitado)
    try:
        github_request("POST", f"/repos/{repo}/pages", token, {
            "source": {"branch": branch, "path": "/"}
        })
        print("  ✅  GitHub Pages habilitado!")
    except urllib.error.HTTPError:
        pass  # já estava habilitado ou sem permissão — ok

    pages_url = f"https://{repo.split('/')[0]}.github.io/{repo.split('/')[1]}/"
    print(f"\n🌐  Dashboard disponível em:")
    print(f"    {pages_url}")
    print("    (Pode levar 1–2 min para o GitHub Pages atualizar)")


# ──────────────────────────────────────────────────────────────────────────────
# 4. MAIN
# ──────────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Atualiza o dashboard de manutenção e publica no GitHub Pages."
    )
    p.add_argument("planilha", help="Caminho para a planilha .xlsx")
    p.add_argument(
        "--dashboard",
        default=os.environ.get("DASHBOARD_FILE", "dashboard_manutencao.html"),
        help="Caminho para o arquivo HTML do dashboard (padrão: dashboard_manutencao.html)",
    )
    p.add_argument(
        "--output",
        default=None,
        help="Salvar HTML atualizado em outro arquivo (padrão: sobrescreve --dashboard)",
    )
    p.add_argument(
        "--repo",
        default=os.environ.get("GITHUB_REPO", ""),
        help="Repositório GitHub: usuario/repo  (ou variável GITHUB_REPO)",
    )
    p.add_argument(
        "--token",
        default=os.environ.get("GITHUB_TOKEN", ""),
        help="Personal Access Token do GitHub (ou variável GITHUB_TOKEN)",
    )
    p.add_argument(
        "--branch",
        default="gh-pages",
        help="Branch do GitHub Pages (padrão: gh-pages)",
    )
    p.add_argument(
        "--remote-file",
        default="index.html",
        help="Nome do arquivo no repositório (padrão: index.html)",
    )
    p.add_argument(
        "--no-publish",
        action="store_true",
        help="Apenas atualiza o HTML localmente, sem publicar no GitHub",
    )
    return p.parse_args()


def main():
    args = parse_args()

    print("=" * 60)
    print("  DASHBOARD DE MANUTENÇÃO – ATUALIZADOR")
    print("=" * 60)

    # 1. Lê a planilha
    machine_data, comp_data = load_xlsx(args.planilha)

    if not machine_data:
        print("❌  Nenhum dado de máquina encontrado. Verifique a planilha.")
        sys.exit(1)

    # 2. Atualiza o HTML
    print(f"\n📝  Atualizando HTML: {args.dashboard}")
    output_path = args.output or args.dashboard
    update_html(args.dashboard, machine_data, comp_data, output_path)

    # 3. Publica no GitHub
    if not args.no_publish:
        if not args.repo:
            print("\n⚠️  --repo não informado. Use --no-publish para pular a publicação.")
            print("   Ex: python atualizar_dashboard.py planilha.xlsx --repo usuario/meu-repo --token ghp_xxx")
            sys.exit(1)
        if not args.token:
            print("\n⚠️  --token não informado. Configure GITHUB_TOKEN ou use --token.")
            sys.exit(1)

        publish_to_github(
            html_path=output_path,
            repo=args.repo,
            token=args.token,
            branch=args.branch,
            remote_filename=args.remote_file,
        )

    print("\n✅  Concluído!")


if __name__ == "__main__":
    main()
